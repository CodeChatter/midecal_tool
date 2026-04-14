"""
批量 OCR 检测脚本：读取 Excel 中的 COS URL，判断图片是否包含有效文字。

支持传入单个 Excel 文件或文件夹（自动处理文件夹下所有 .xlsx/.xls 文件）。

用法:
    直接修改下方 ════ 配置区 ════ 的变量，然后运行:
    python -m tests.test_ocr_check
"""

import logging
import os
import queue
import sys
import tempfile
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import List, Optional, Tuple

import openpyxl

# 将项目根目录加入 path，以便直接运行
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv

load_dotenv(ROOT / ".env")

from app.core.ocr.paddle import PaddleOCREngine
from app.core.models import TextLine
from app.core.cos import parse_url, download_to_local

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)

# ════════════════════════════════════════════════════════════
#  配置区 —— 修改以下变量即可，无需传命令行参数
# ════════════════════════════════════════════════════════════

# 输入路径：Excel 文件 或 包含 Excel 的文件夹
INPUT_PATH = "/Users/chatter/development/devBank/Python/flask/demo1/cos_output/"

# COS URL 所在列名
COL_NAME = "文件URL"

# 最少字符数阈值，低于此值视为"无文字"
CHAR_THRESHOLD = 5

# 并发下载线程数
WORKERS = 8

# OCR 引擎并发数（每个引擎约占 500MB 内存，按机器内存调整）
OCR_WORKERS = 3

# 输出目录（空字符串 = 与输入同目录，自动加 _result 后缀）
OUTPUT_DIR = ""

# ════════════════════════════════════════════════════════════

# ─── OCR 引擎池（多实例并行） ───
_ocr_pool: Optional[queue.Queue] = None


def init_ocr_pool(n: int):
    """初始化 N 个 OCR 引擎实例，放入队列供并发获取。"""
    global _ocr_pool
    _ocr_pool = queue.Queue()
    for i in range(n):
        print(f"  初始化 OCR 引擎 [{i + 1}/{n}]...", flush=True)
        engine = PaddleOCREngine()
        _ocr_pool.put(engine)
    print(f"  OCR 引擎池就绪 ({n} 个实例)", flush=True)


def download_image(url: str) -> Optional[str]:
    """通过 COS SDK 下载图片到临时文件（支持私有 bucket），返回本地路径。"""
    try:
        from urllib.parse import urlparse
        parsed = urlparse(url)
        suffix = Path(parsed.path).suffix or ".jpg"
        fd, local_path = tempfile.mkstemp(suffix=suffix)
        os.close(fd)

        loc = parse_url(url)
        download_to_local(loc, local_path)
        return local_path
    except Exception as e:
        logger.error(f"下载失败 {url}: {e}")
        return None


def ocr_check(local_path: str, min_chars: int) -> Tuple[int, int, str, bool]:
    """
    对单张图片执行 OCR，返回 (文本行数, 总字符数, 拼接文本, 是否有效文字)。
    从引擎池借用一个引擎，用完归还，实现真正并行 OCR。
    """
    engine = _ocr_pool.get()
    try:
        lines: List[TextLine] = engine.recognize(local_path)
    finally:
        _ocr_pool.put(engine)

    total_text = "".join(line.text for line in lines)
    char_count = len(total_text.replace(" ", "").replace("\n", ""))
    has_text = char_count >= min_chars

    return len(lines), char_count, total_text.strip(), has_text


def process_one(
        idx: int, url: str, min_chars: int
) -> dict:
    """处理单条记录：下载 → OCR → 判断 → 清理临时文件。"""
    start = time.time()
    result = {
        "index": idx,
        "url": url,
        "line_count": 0,
        "char_count": 0,
        "has_text": False,
        "ocr_text": "",
        "error": "",
        "elapsed_s": 0.0,
    }

    local_path = download_image(url)
    if not local_path:
        result["error"] = "下载失败"
        result["elapsed_s"] = round(time.time() - start, 2)
        return result

    try:
        line_count, char_count, ocr_text, has_text = ocr_check(local_path, min_chars)
        result.update(
            line_count=line_count,
            char_count=char_count,
            has_text=has_text,
            ocr_text=ocr_text[:200],  # 截断，避免过长
        )
    except Exception as e:
        result["error"] = str(e)
    finally:
        if os.path.exists(local_path):
            os.unlink(local_path)

    result["elapsed_s"] = round(time.time() - start, 2)
    return result


def read_urls_from_excel(filepath: str, col_name: str) -> Tuple[List[str], List[Tuple[int, str, List]]]:
    """
    从 Excel 读取 URL 列，返回 (表头列表, [(行号, url, 整行数据), ...])。
    保留整行数据用于后续输出 _have_text 文件。
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    # 找到目标列
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        col_idx = header.index(col_name)
    except ValueError:
        raise ValueError(
            f"找不到列 '{col_name}'，可用列: {header}"
        )

    rows = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        row_values = [cell.value for cell in row]
        val = row_values[col_idx] if col_idx < len(row_values) else None
        if val and isinstance(val, str) and val.strip().startswith("http"):
            rows.append((row_num, val.strip(), row_values))

    wb.close()
    return header, rows


def write_results(results: List[dict], output_path: str):
    """将结果写入 Excel。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OCR检测结果"

    headers = [
        "行号", "COS URL", "OCR行数", "字符数",
        "有效文字", "OCR文本(前200)", "错误", "耗时(s)",
    ]
    ws.append(headers)

    for r in sorted(results, key=lambda x: x["index"]):
        ws.append([
            r["index"],
            r["url"],
            r["line_count"],
            r["char_count"],
            "是" if r["has_text"] else "否",
            r["ocr_text"],
            r["error"],
            r["elapsed_s"],
        ])

    wb.save(output_path)
    logger.info(f"结果已保存: {output_path}")


def write_have_text_excel(
        header: List[str],
        rows_with_text: List[List],
        output_path: str,
):
    """将有文字的行按原始格式输出到新 Excel。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "有文字图片"

    ws.append(header)
    for row_values in rows_with_text:
        ws.append(row_values)

    wb.save(output_path)
    logger.info(f"有文字结果已保存: {output_path} ({len(rows_with_text)} 条)")


def collect_excel_files(input_path: Path) -> List[Path]:
    """收集需要处理的 Excel 文件列表。支持单文件或文件夹。"""
    if input_path.is_file():
        return [input_path]

    if input_path.is_dir():
        files = sorted(
            f for f in input_path.iterdir()
            if f.suffix.lower() in (".xlsx", ".xls") and not f.name.startswith("~")
        )
        return files

    return []


def process_single_excel(
        excel_path: Path,
        col_name: str,
        threshold: int,
        workers: int,
        result_path: str,
        have_text_path: str,
) -> dict:
    """处理单个 Excel 文件，返回汇总统计。"""
    # 1. 读取 URL（同时保留原始行数据）
    header, rows = read_urls_from_excel(str(excel_path), col_name)
    total = len(rows)

    summary = {
        "file": excel_path.name,
        "total": total,
        "has_text": 0,
        "no_text": 0,
        "error": 0,
        "elapsed_s": 0.0,
    }

    if total == 0:
        logger.warning(f"  文件 {excel_path.name} 没有有效 URL，跳过")
        return summary

    # 建立 行号 → 原始行数据 的映射
    row_data_map = {row_num: row_values for row_num, _, row_values in rows}
    url_list = [(row_num, url) for row_num, url, _ in rows]

    # 2. 并发处理
    results: List[dict] = []
    done_count = 0
    start_all = time.time()

    def _progress():
        pct = int(done_count / total * 100)
        bar_len = 30
        filled = int(bar_len * done_count / total)
        bar = "█" * filled + "░" * (bar_len - filled)
        elapsed = time.time() - start_all
        speed = done_count / elapsed if elapsed > 0 else 0
        eta = (total - done_count) / speed if speed > 0 else 0
        print(
            f"\r  {bar} {pct:3d}% ({done_count}/{total}) "
            f"有文字:{summary['has_text']} 无文字:{summary['no_text']} 错误:{summary['error']} "
            f"[{elapsed:.0f}s<{eta:.0f}s, {speed:.1f}张/s]",
            end="", flush=True,
        )

    _progress()  # 显示初始状态

    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {
            pool.submit(process_one, idx, url, threshold): (idx, url)
            for idx, url in url_list
        }

        for future in as_completed(futures):
            r = future.result()
            results.append(r)
            done_count += 1

            if r["error"]:
                summary["error"] += 1
            elif r["has_text"]:
                summary["has_text"] += 1
            else:
                summary["no_text"] += 1

            _progress()

    print(flush=True)  # 换行

    summary["elapsed_s"] = round(time.time() - start_all, 1)

    # 3. 输出完整检测结果
    write_results(results, result_path)

    # 4. 输出有文字的行（保持原始 Excel 格式）
    has_text_rows = [
        row_data_map[r["index"]]
        for r in sorted(results, key=lambda x: x["index"])
        if r["has_text"]
    ]
    if has_text_rows:
        write_have_text_excel(header, has_text_rows, have_text_path)

    logger.info(f"  完成: 有文字={summary['has_text']}, 无文字={summary['no_text']}, "
                f"错误={summary['error']}, 耗时={summary['elapsed_s']}s")

    return summary


def main():
    input_path = Path(INPUT_PATH)
    if not input_path.exists():
        logger.error(f"输入路径不存在: {input_path}")
        sys.exit(1)

    # 收集 Excel 文件
    excel_files = collect_excel_files(input_path)
    if not excel_files:
        logger.error(f"未找到 Excel 文件: {input_path}")
        sys.exit(1)

    logger.info(f"共发现 {len(excel_files)} 个 Excel 文件:")
    for i, f in enumerate(excel_files, 1):
        logger.info(f"  [{i}] {f.name}")
    logger.info(f"配置: 阈值={CHAR_THRESHOLD}字, 并发={WORKERS}线程")

    # 确定输出目录
    if OUTPUT_DIR:
        out_dir = Path(OUTPUT_DIR)
        out_dir.mkdir(parents=True, exist_ok=True)
    elif input_path.is_dir():
        out_dir = input_path
    else:
        out_dir = input_path.parent

    # 初始化 OCR 引擎池（多实例并行，首次加载模型较慢）
    print(f"\n初始化 OCR 引擎池 ({OCR_WORKERS} 个实例)...", flush=True)
    init_ocr_pool(OCR_WORKERS)

    # 逐个处理 Excel 文件
    all_summaries: List[dict] = []
    total_files = len(excel_files)
    total_start = time.time()

    for file_idx, excel_file in enumerate(excel_files, 1):
        logger.info(f"\n[{file_idx}/{total_files}] {excel_file.name}")
        result_path = str(out_dir / f"{excel_file.stem}_result.xlsx")
        have_text_path = str(out_dir / f"{excel_file.stem}_have_text.xlsx")
        summary = process_single_excel(
            excel_file, COL_NAME, CHAR_THRESHOLD, WORKERS, result_path, have_text_path,
        )
        all_summaries.append(summary)

    total_elapsed = round(time.time() - total_start, 1)

    # 全局汇总
    grand_total = sum(s["total"] for s in all_summaries)
    grand_has = sum(s["has_text"] for s in all_summaries)
    grand_no = sum(s["no_text"] for s in all_summaries)
    grand_err = sum(s["error"] for s in all_summaries)

    logger.info(f"\n{'=' * 60}")
    logger.info(f"全部完成 | {total_files}个文件 | {grand_total}张图片 | 耗时{total_elapsed}s")
    logger.info(f"有文字: {grand_has} | 无文字: {grand_no} | 错误: {grand_err}")
    logger.info(f"{'=' * 60}")

    for s in all_summaries:
        logger.info(
            f"  {s['file']}: {s['total']}张 → "
            f"有文字={s['has_text']}, 无文字={s['no_text']}, "
            f"错误={s['error']}, {s['elapsed_s']}s"
        )


if __name__ == "__main__":
    main()
